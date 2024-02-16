from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait 
from selenium.webdriver.support import expected_conditions as ec 
import pytest
import openpyxl
from constants import global_constants as c

class TestSauceIde:
    def setup(self): 
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window() 

    def teardown(self): 
        self.driver.quit()
    
    def get_data_invalid_password():
        excel = openpyxl.load_workbook(c.XLSX)
        sheet = excel["Sayfa1"] #hangi sayfada çalışacağımı gösteriyorum
        data = []
        for i in range(2,3): #3.satırdan 4. satıra kadar
            username = sheet.cell(i,1).value #1.sütun
            password = sheet.cell(i,2).value #2.sütun
            data.append((username,password))
        return data
       

     
    def get_data_invalid_lock():
        excel = openpyxl.load_workbook(c.XLSX)
        sheet = excel["Sayfa1"]  
        data = []
        for i in range(3,4): 
            username = sheet.cell(i,1).value 
            password = sheet.cell(i,2).value 
            data.append((username,password))
        return data



    def get_data_valid_login():
        excel = openpyxl.load_workbook(c.XLSX)
        sheet = excel["Sayfa1"] 
        rows = sheet.max_row #kaçıncı satıra kadar veri var?
        data = []
        for i in range(4,5): 
            username = sheet.cell(i,1).value 
            password = sheet.cell(i,2).value 
            data.append((username,password))
        return data 

    

    def get_data_parameter():
        excel = openpyxl.load_workbook(c.XLSX)
        sheet = excel["Sayfa1"] 
        rows = sheet.max_row
        data = []
        for i in range(5,rows+1): 
            username = sheet.cell(i,1).value 
            password = sheet.cell(i,2).value 
            data.append((username,password))
        return data
    


    def test_invalid_user(self):
        self.driver.find_element(By.ID,c.LOGIN_BUTTON).click()
        error_message_user = self.driver.find_element(By.XPATH,c.USER_MESSAGE_XPATH)
        assert error_message_user.text == c.USER_MESSAGE

    
    @pytest.mark.parametrize("username,password",get_data_invalid_password())
    def test_invalid_password(self,username,password):
        WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME))).send_keys(username)
        WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD))).send_keys(password)
        self.driver.find_element(By.ID,c.LOGIN_BUTTON).click()
        error_message_password = self.driver.find_element(By.XPATH,c.PASSWORD_MESSAGE_XPATH)
        assert error_message_password.text == c.PASSWORD_MESSAGE

    @pytest.mark.parametrize("username,password",get_data_invalid_lock())
    def test_invalid_lock(self,username,password):
        WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME))).send_keys(username)
        WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD))).send_keys(password)
        self.driver.find_element(By.ID,c.LOGIN_BUTTON).click()
        error_message_lock = self.driver.find_element(By.XPATH,c.LOCK_MESSAGE_XPATH)
        assert error_message_lock.text == c.LOCK_MESSAGE
    
    
    @pytest.mark.parametrize("username,password",get_data_valid_login())
    def test_valid_item(self,username,password):
        WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME))).send_keys(username)
        WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD))).send_keys(password)
        self.driver.find_element(By.ID,c.LOGIN_BUTTON).click()
        inventory_items = self.driver.find_elements(By.CLASS_NAME,c.ITEM_NAME)
        assert len(inventory_items) == c.LEN


    @pytest.mark.parametrize("username,password",get_data_parameter()) 
    def test_invalid_match(self,username,password):
        WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME))).send_keys(username)
        WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD))).send_keys(password)
        self.driver.find_element(By.ID,c.LOGIN_BUTTON).click()
        errorMessage = self.driver.find_element(By.XPATH,c.MATCH_MESSAGE_XPATH)
        assert errorMessage.text == c.MATCH_MESSAGE


    @pytest.mark.parametrize("username,password",get_data_valid_login()) 
    def test_add_product(self,username,password):
        WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME))).send_keys(username)
        WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD))).send_keys(password)
        self.driver.find_element(By.ID,c.LOGIN_BUTTON).click()
        WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.ADDTOCART_XPATH))).click()
        WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.SHOPPINGCART_LINK_XPATH))).click()
        product = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.PRODUCT_XPATH)))
        assert product.text == c.PRODUCT_NAME
        WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.CONTINUE_SHOPPING_XPATH))).click()
        remove = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.REMOVE_XPATH)))
        assert remove.text == c.REMOVE

    
    @pytest.mark.parametrize("username,password",get_data_valid_login())  
    def test_product_review(self,username,password):
        WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME))).send_keys(username)
        WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD))).send_keys(password)
        self.driver.find_element(By.ID,c.LOGIN_BUTTON).click()
        self.driver.execute_script("window.scrollTo(0,500)")
        WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.ITEM_NAME_XPATH ))).click()
        product_title = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.PRODUCT_TITLE_XPATH)))
        assert product_title.text == c.PRODUCT_TITLE
        self.driver.find_element(By.XPATH,c.BACK_BUTTON_XPATH).click()